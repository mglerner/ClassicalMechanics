"""Generate the Fall 2026 PHY 317 (Classical Mechanics, Smith) course calendar.

Layout (2026-09-02, mirroring PHY 210's generator): a single-block
"Schedule" sheet with
Week | Class | Date | Topics | Reading Due | Prerequisites | PCCI | HW Due | Exams,
written as two chunks (header row repeated) split at fall break for the
half-semester Moodle embed, plus "HW Problem Lists" and "Grade Categories"
sheets. The old two-side-by-side-blocks print layout is in git history
before this date.

Content follows Will Raven's Fall 2025 PHY 317 sequence (Taylor,
Classical Mechanics) remapped onto the Smith Fall 2026 academic calendar;
his 39 teaching slots (including the Mountain Day holder and catch-up
days) map one-to-one onto Fall 2026's 39 MWF meetings. His homework sets
are kept intact (HW Problem Lists sheet); only the due dates move.

Usage: python make_fall2026_calendar.py OUTPUT.xlsx
"""
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------- semester
# Smith Fall 2026: classes Tue Sep 8 - Tue Dec 15.
# MWF meetings; skip Mon Oct 12 (autumn recess), Wed Nov 25 + Fri Nov 27
# (Thanksgiving). Cromwell Day (Tue Nov 10) doesn't hit MWF.
# PHY 317 meets W/F 1:20-2:35 and Mon 1:40-2:55 (registrar, verified
# 2026-08-17) -- same MWF days, so the 39-meeting mapping holds.
FIRST_DAY = date(2026, 9, 9)   # first MWF meeting (classes open Tue Sep 8)
LAST_DAY = date(2026, 12, 14)  # last MWF meeting
NO_CLASS = {
    date(2026, 10, 12): "Autumn recess",
    date(2026, 11, 25): "Thanksgiving",
    date(2026, 11, 27): "Thanksgiving",
}

def class_days():
    d = FIRST_DAY
    while d <= LAST_DAY:
        if d.weekday() in (0, 2, 4) and d not in NO_CLASS:  # M W F
            yield d
        d += timedelta(days=1)

# ------------------------------------------------------------------ content
# (topic, reading, prerequisites) per teaching slot, in Will Raven's F2025
# order. Readings are Taylor chapter/sections, read before class.
# Prerequisites are Will's just-in-time review pointers: Knight = intro
# physics (PHY 117/118), Math Methods = Felder & Felder (PHY 210).
# Exams are pinned to slot indices (0-based). Decided 2026-09-04 (Michael):
# exams as early as the homework allows -- each exam comes after the last
# homework on its chapters has been RETURNED.
#   Exam 1 = index 13 = Fri Oct 9: Will's exact slot (his slot 14 of 39,
#     the class after 5.1-5.2). HW04 (Ch 4) is due Mon Oct 5 and comes back
#     Wed Oct 7.
#   Exam 2 = index 26 = Wed Nov 11 (Will's was 27). HW09 (the Ch 7
#     practice) is due Fri Nov 6 and comes back Mon Nov 9, the review day.
#   Exam 3 = index 35 = Mon Dec 7 (Will's was 37; Michael: fine as is).
# Every slot is 75 min, so the weekday changes nothing else.
# Exam coverage per Will's syllabus (his calendar row "Ch 9-11" for Exam 3
# was a typo; the 3x3 redemption-problem arithmetic proves 8, 9, 11).
EXAMS = {
    13: ("EXAM 1 (Ch 2-4)", {2, 3, 4}),          # Fri Oct 9
    26: ("EXAM 2 (Ch 5-7)", {5, 6, 7}),          # Wed Nov 11
    35: ("EXAM 3 (Ch 8, 9, 11)", {8, 9, 11}),    # Mon Dec 7
}
CONTENT = [
    ("Syllabus; notation; Newton's laws; polar coordinates", "Ch. 1",
     "Knight: Ch. 3 (Vectors), Ch. 4-6 (Forces)"),
    ("Newton's laws", "Ch. 1",
     "Knight: Ch. 3 (Vectors), Ch. 4-6 (Forces)\n"
     "Math Methods: Ch. 5 (Cylindrical and Spherical Coordinates)\n"
     "Other: Algebra"),
    ("Linear drag force; terminal velocity", "2.1-2.2",
     "Knight: Ch. 5 (Drag)\nMath Methods: Ch. 1 (Intro to ODEs)\n"
     "Other: Algebra, graphs, integrals, derivatives"),
    ("Trajectories and range; quadratic air drag", "2.3-2.4",
     "Knight: Ch. 4 (Kinematics in 2D)\nMath Methods: Ch. 1 (Intro to ODEs)\n"
     "Other: Algebra, graphs, integrals, derivatives"),
    # Reading widened 2.5 -> 2.5-2.7 (coverage audit 2026-09-03): Will's
    # labels were narrower than what he taught; HW02's 2.53 is section 2.7.
    ("Lorentz force law; cyclotron motion", "2.5-2.7",
     "Knight: Ch. 3 (Vectors), Ch. 22-23 (Electromagnetism Review)\n"
     "Math Methods: Ch. 3 (Complex numbers, Euler's equation), "
     "Ch. 8 (Vector Calculus)\nOther: Cross products"),
    ("Conservation of momentum; rocket motion", "3.1-3.2",
     "Knight: Ch. 11 (Impulse and Momentum)\n"
     "Other: Integrals, vector components"),
    ("Center of mass; angular momentum", "3.3-3.4",
     "Knight: Ch. 11 (Momentum), Ch. 12 (Rotation)\n"
     "Other: Integrals, vector components, cross products"),
    ("Moment of inertia; conservation of angular momentum", "3.5",
     "Knight: Ch. 12 (Rotation of Rigid Body)\n"
     "Other: Integrals, vector components"),
    ("Work-KE theorem; conservative forces", "4.1-4.3",
     "Knight: Ch. 9-10 (Work and Kinetic Energy)\n"
     "Math Methods: Ch. 8 (Line integrals, curls)"),
    ("Potential energy; graphs of PE functions", "4.4-4.7",
     "Knight: Ch. 9-10 (Potential Energy)\n"
     "Other: Full derivatives and chain rule"),
    ("Central forces; multiparticle systems", "4.8-4.10",
     "Knight: Ch. 12 (Gravity)\nMath Methods: Ch. 5 (Spherical "
     "Coordinates), Ch. 8 (Gradients in spherical)"),
    ("Mountain Day holder (if Mountain Day has already happened: "
     "Ch 1-4 problem session)", "", ""),
    ("Spring forces; simple harmonic oscillator", "5.1-5.2",
     "Knight: Ch. 15 (Oscillations)\nMath Methods: Ch. 6 (Oscillations)"),
    ("2D oscillators; damped SHO", "5.3-5.4",
     "Math Methods: Ch. 1 (ODEs)"),
    ("Forced damped SHO; resonance", "5.5-5.6",
     "Math Methods: Ch. 3 (Complex Numbers), Ch. 1 (ODEs)"),
    ("Fourier series analysis of driven SHO", "5.7-5.8",
     "Math Methods: Ch. 9 (Fourier Series and Transforms)"),
    ("Calculus of variations; Fermat's principle; Euler-Lagrange", "6.1-6.2",
     "Math Methods: Ch. 4 (Partial Derivatives), Ch. 5 (Integration)"),
    ("Euler-Lagrange equation; brachistochrone problem", "6.3",
     "Math Methods: Ch. 4 (Partial Derivatives)"),
    ("Multiple variables; generalized coordinates", "6.4", ""),
    ("Lagrange's equations; Hamilton's principle", "7.1", ""),
    ("Constrained systems; generalized coordinates", "7.2-7.3", ""),
    # 7.8 added 2026-09-02 (Michael asked for a simple Noether's theorem):
    # Taylor 7.8 IS the simple version -- translation invariance of L gives
    # conservation of P, time-translation invariance gives conservation of
    # the Hamiltonian, and Taylor names Noether explicitly. Will never
    # assigned 7.8 (his only Noether mention was one line of Ch 4 ink).
    # Will's row read "7.5, 7.9" with the topic "Lagrange multipliers";
    # in Taylor, multipliers are 7.10 and 7.9 is magnetic forces. The topic
    # is what he taught (his HW08 has 7.10-the-problem, none from 7.9), so
    # the reading says 7.10. Flagged in TODO.
    ("Examples of Lagrange's equations; Noether's theorem (symmetries and "
     "conservation laws); Lagrange multipliers", "7.5, 7.8, 7.10",
     ""),
    ("Central forces; reduced mass; equations of motion", "8.1-8.4",
     "Knight: Ch. 12 (Gravity)"),
    ("Orbits", "8.5-8.6",
     "Math Methods: Ch. 1 (ODEs)\nCalculus I & II"),
    # Review day BEFORE Exam 2 (Wed); Ch 8 is not on Exam 2, so changing
    # orbits can follow the exam.
    ("Catch-up / review day (Exam 2 is Wednesday)", "", ""),
    ("Changing orbits", "8.7-8.8", "Calculus I & II"),
    ("Accelerating frames; tides", "9.1-9.2",
     "Math Methods: Ch. 8 (Vector Calculus)"),
    ("Angular velocity; rotating frames", "9.3-9.5",
     "Math Methods: Ch. 8 (Vectors in Curvilinear Coordinates)"),
    # 9.6-9.7 -> 9.6-9.8 (audit 2026-09-03): HW11's 9.28, 9.29 are 9.8.
    ("Centrifugal force; Coriolis force", "9.6-9.8", ""),
    ("Coupled oscillators; two masses and three springs", "11.1-11.2",
     "Knight: Ch. 15 (Oscillations)\nMath Methods: Ch. 6 "
     "(Eigenvectors/Eigenvalues)\nReview the ideas behind basis vectors"),
    ("Normal coordinates; weakly coupled oscillators; start the double "
     "pendulum", "11.2-11.4",
     "Math Methods: Ch. 6 (Eigenvectors/Eigenvalues)"),
    # Swapped 2026-09-02 (Will had catch-up then double pendulum): the
    # double pendulum finishes Ch 11 two days before HW12 is due, and the
    # catch-up day becomes the review day right before Exam 3.
    ("Double pendulum", "11.4", ""),
    ("Catch-up / review day (Exam 3 is Monday)", "", ""),
    # 12.1-12.3 -> 12.1-12.5 (audit 2026-09-03): Will's chaos notebooks are
    # bifurcation and Lyapunov exponents = 12.4-12.5; HW13 draws on both.
    ("Chaos: the driven damped pendulum; period doubling", "12.1-12.4", ""),
    # Liouville aside lives here (TODO item 20; materials in
    # ../../../FluctuationTheorems/01-classical-mechanics/).
    ("Chaos: sensitivity to initial conditions; Liouville aside",
     "12.4-12.5", ""),
    ("Last day: review and wrap-up", "", ""),
]

# -------------------------------------------------------------------- HW
# Will's 13 sets are kept intact; only the due dates move. His HW was due
# Fridays covering the Wed-Fri-Mon block that ended four days earlier. Our
# semester opens on a Wednesday, so the same blocks fall Mon-Wed-Fri and a
# Friday deadline would land ON the block's last teaching day (the exact
# defect PHY 210's coverage audit found). Rule adopted 2026-09-02:
#
#   HW is due WEDNESDAYS at 10:00 PM and covers the previous Mon/Wed/Fri
#   (five days after the last class it draws on), with these one-offs:
#   HW04 Mon Oct 5 (so it is back before Exam 1 on Fri Oct 9); HW06 Fri
#   Oct 23 (5.7-5.8 is taught Mon Oct 19); HW09 Fri Nov 6 (so it is back
#   Mon Nov 9, before Exam 2 on Wed Nov 11); HW11 Mon Nov 23 (the
#   Wednesday is Thanksgiving; Will made the same Monday move); HW12 Fri
#   Dec 4 (the Ch 11 practice must precede Exam 3). HW13 lands in reading
#   period (Will's was the day after his last class).
#
# Tuples: (hw, due, covers_through, chapters, covers, problems).
# covers_through = the last class day whose material the set draws on;
# the build asserts due > covers_through, that no due date lands on a
# no-class day or an exam day, and that every set precedes the exam that
# covers its chapters. Problem lists from Will's F2025 PDFs (private/
# WillF2025/MoodleCourse/extracted/03_Homework/); custom problems are
# written out in full in the Moodle descriptions, not here.
HWS = [
    (1, date(2026, 9, 16), date(2026, 9, 11), {1},
     "Ch 1: Newton's laws, polar coordinates",
     "1.27, 1.45, 1.46, 1.49 [see in-class 1.47 for cylindrical "
     "coordinates], 1.50 [computational]"),
    (2, date(2026, 9, 23), date(2026, 9, 18), {2},
     "Ch 2: projectiles with drag; charged particles",
     "2.14 [look up the integral], 2.31, 2.36, 2.39, 2.42 [2.41 done in "
     "class], 2.53"),
    (3, date(2026, 9, 30), date(2026, 9, 25), {3},
     "Ch 3: momentum, rockets, center of mass, angular momentum",
     "3.8, 3.13 [uses 3.11(b) from class], 3.19, 3.35, 3.36; challenge: "
     "chain pulled off a table at constant speed"),
    (4, date(2026, 10, 5), date(2026, 10, 2), {4},
     "Ch 4: energy",
     "4.4, 4.8, 4.13 [optional math review], 4.24(a-c), 4.36, 4.39 "
     "[advanced, optional], 4.53; custom: variation of gravity with height"),
    (5, date(2026, 10, 14), date(2026, 10, 7), {5},
     "5.1-5.2: Hooke's law, simple harmonic motion",
     "5.4 [geometry diagram supplied], 5.7, 5.13"),
    (6, date(2026, 10, 23), date(2026, 10, 19), {5},
     "5.3-5.8: 2D and damped oscillators, resonance, Fourier series",
     "5.28, 5.32 [optional], 5.43; custom: RLC circuit as an SHO; custom: "
     "square-wave Fourier drive (or 5.49)"),
    (7, date(2026, 10, 28), date(2026, 10, 23), {6},
     "6.1-6.3: calculus of variations, Euler-Lagrange, brachistochrone",
     "6.7, 6.11, 6.14, 6.17; custom: brachistochrone revisited (6.25 "
     "recast; tautochrone)"),
    (8, date(2026, 11, 4), date(2026, 10, 30), {6, 7},
     "6.4, 7.1-7.3: several variables; Lagrange's equations; constraints",
     "6.23 [hints], 6.24 [advanced, optional], 6.27, 7.3, 7.10 [hint], 7.14"),
    # Fri, not Wed: back on Mon Nov 9 before Exam 2. 8.2 (taught Wed Nov 4)
    # gets only two days; it is the one easy problem on the set.
    (9, date(2026, 11, 6), date(2026, 11, 4), {7, 8},
     "7.5, 7.8, 7.9, 8.1-8.2: Lagrangian examples; Noether; multipliers; "
     "two-body setup",
     "7.27, 7.31, 7.34, 7.38 [optional challenge], 7.46 [Noether: rotational "
     "invariance gives conservation of L_z], 8.2"),
    (10, date(2026, 11, 18), date(2026, 11, 13), {8},
     "8.3-8.8: central-force motion, orbits, changing orbits",
     "8.9, 8.13, 8.14 [optional challenge], 8.15, 8.18, 8.23"),
    (11, date(2026, 11, 23), date(2026, 11, 20), {8, 9},
     "8.8, Ch 9: non-inertial frames, Coriolis",
     "8.27 [optional challenge], 8.33, 9.3, 9.9, 9.15 [Northampton "
     "colatitude 48 deg], 9.17, 9.28, 9.29 [optional challenge]"),
    (12, date(2026, 12, 4), date(2026, 12, 2), {11},
     "Ch 11: coupled oscillators, normal modes, double pendulum",
     "11.6, 11.9, 11.12, 11.14 [Lagrangian supplied as a check], 11.17 "
     "[optional]"),
    (13, date(2026, 12, 16), date(2026, 12, 11), {12},
     "Ch 12: chaos",
     "12.1, 12.6, 12.7 [computational: use the class program or write your "
     "own], 12.8-12.10 [optional], 12.13 + custom parts (b), (c)"),
]

# Extra (non-HW) due dates shown in the HW Due column.
EXTRA_DUE = {}

# ---------------------------------------------------------------- PCCIs
# Pre-Class Check-Ins, same machinery as PHY 210 (decision 2026-09-02):
# a short problem due on paper at the start of class, good-faith graded,
# the participation/attendance artifact. Will's "Look At" problems (the
# Taylor problems he had students look at before class; solutions posted
# per chapter, transcribed 2026-09-02 from his image PDFs into
# private/F2026PrepPacks/_shared/will-problem-tables.md) are the source
# wherever he had one for that day; where his day had none, a one-line
# reading prompt. One problem per day (Taylor's easiest-starred where he
# listed two) so the <15-minute promise holds. Exam days get none.
PCCI = {
    # No PCCI on day 1 (nobody has the syllabus before the first class).
    # Ch 1 (Will's Look-At, no day split: 1.4, 1.6, 1.10, 1.11, 1.31, 1.35)
    date(2026, 9, 11): "Read the syllabus and bring one question or comment; "
                       "and 1.10",
    # Ch 2 (Will's Look-At: Wed 2.1, 2.5, 2.7 / Fri 2.16, 2.25 / Mon 2.49;
    # he circled 2.7 in class)
    date(2026, 9, 14): "2.7",
    date(2026, 9, 16): "2.25",
    date(2026, 9, 18): "2.49",
    # Ch 3 (Will's Look-At: Wed 3.7 / Fri 3.16, 3.17, 3.25 / Mon 3.32)
    date(2026, 9, 21): "3.7",
    date(2026, 9, 23): "3.16",
    date(2026, 9, 25): "3.32",
    # Ch 4 (Will's Look-At: Wed 4.7, 4.16 / Fri 4.31 / Mon 4.41)
    date(2026, 9, 28): "4.7",
    date(2026, 9, 30): "4.31",
    date(2026, 10, 2): "4.41",
    date(2026, 10, 5): "Bring one Ch 1-4 problem you want worked in the "
                       "problem session (skip if Mountain Day)",
    # Ch 5
    date(2026, 10, 7): "5.1",
    date(2026, 10, 14): "5.21",
    date(2026, 10, 16): "5.35",
    date(2026, 10, 19): "5.47",
    # Ch 6-7
    date(2026, 10, 21): "6.3",
    date(2026, 10, 23): "6.8",
    date(2026, 10, 26): "6.20",
    date(2026, 10, 28): "7.1",
    date(2026, 10, 30): "Read 7.2-7.3: in a sentence, what makes a "
                        "constraint holonomic? Give one example of each kind",
    date(2026, 11, 2): "7.17",
    # Ch 8
    date(2026, 11, 4): "8.6",
    date(2026, 11, 6): "8.19",
    date(2026, 11, 9): "Bring one Ch 5-7 problem you want worked in review",
    date(2026, 11, 13): "8.28",
    # Ch 9
    date(2026, 11, 16): "9.1",
    date(2026, 11, 18): "9.7",
    date(2026, 11, 20): "9.12",
    # Ch 11
    date(2026, 11, 23): "11.1",
    date(2026, 11, 30): "Read 11.3: in a sentence or two, what happens to "
                        "the energy of two weakly coupled pendulums over time?",
    date(2026, 12, 2): "Write down the Lagrangian of a simple pendulum "
                       "(length L, angle phi); no need to solve it",
    date(2026, 12, 4): "Bring one Ch 8, 9, or 11 problem you want worked "
                       "in review",
    # Ch 12
    date(2026, 12, 9): "Read 12.1-12.2: in a sentence, what does it mean "
                       "for an equation of motion to be linear, and what do "
                       "we lose when it isn't?",
    date(2026, 12, 11): "Log into posit.smith.edu and open the driven-"
                        "pendulum notebook on Moodle; bring your laptop",
    date(2026, 12, 14): "Bring one question for the final-exam review",
}


# ------------------------------------------------- chapter problem lists
# Will's three streams per chapter, transcribed from his F2025 Moodle
# backup (private/F2026PrepPacks/_shared/will-problem-tables.md and the
# Ch 1-4 deck extractions): "Look at" = problems students study before
# class (solutions posted; OUR PCCIs are drawn from these, one per day),
# "In class" = worked together in class, "Homework" = the graded sets
# (kept exactly as Will assigned them, plus 7.46 on HW09). Days are
# class days within the chapter in our sequence. This feeds the
# "Chapter Problems" sheet and ChapterProblemLists.md, the source for the
# chapter-opening slide (Look at / In class / Homework, three columns).
CHAPTER_PROBLEMS = {
    1: dict(lookat=[["1.4", "1.6", "1.10", "1.11", "1.31", "1.35"]],
            inclass=[["1.9", "1.18", "1.26", "1.43", "1.47", "1.48"]],
            hw=[("HW01", ["1.27", "1.45", "1.46", "1.49", "1.50 (computational)"])]),
    2: dict(lookat=[["2.1", "2.5", "2.7"], ["2.16", "2.25"], ["2.49"]],
            inclass=[["2.4", "2.11", "2.12", "2.13"], ["2.33", "2.34", "2.35", "2.41"],
                     ["2.52", "2.54"]],
            hw=[("HW02", ["2.14", "2.31", "2.36", "2.39", "2.42", "2.53"])]),
    3: dict(lookat=[["3.7"], ["3.16", "3.17", "3.25"], ["3.32"]],
            inclass=[["3.5", "3.10", "3.11"], ["3.21", "3.27"], ["3.29", "3.34"]],
            hw=[("HW03", ["3.8", "3.13", "3.19", "3.35", "3.36",
                          "Challenge: chain lifted at constant speed"])]),
    4: dict(lookat=[["4.7", "4.16"], ["4.31"], ["4.41"]],
            inclass=[["4.2", "4.3", "4.12", "4.15"], ["4.23", "4.26", "4.28"],
                     ["4.46", "4.48"]],
            hw=[("HW04", ["4.4", "4.8", "4.13 (optional)", "4.24(a-c)",
                          "Variation of gravity with height", "4.36",
                          "4.39 (advanced, optional)", "4.53"])]),
    5: dict(lookat=[["5.1", "5.3"], ["5.21"], ["5.35"], ["5.47"]],
            inclass=[["5.6", "5.8", "5.12"], ["5.22", "5.26"],
                     ["5.40", "5.41", "5.42", "5.44"], ["5.49"]],
            hw=[("HW05", ["5.4", "5.7", "5.13"]),
                ("HW06", ["5.28", "5.32 (optional)", "5.43", "RLC circuit as an SHO",
                          "Square-wave drive (or 5.49)"])]),
    6: dict(lookat=[["6.3", "6.4"], ["6.8"], ["6.20"]],
            inclass=[["6.1", "6.2"], ["6.9", "6.18"], ["6.21", "6.25"]],
            hw=[("HW07", ["6.7", "6.11", "6.14", "6.17", "Brachistochrone revisited"]),
                ("HW08", ["6.23", "6.24 (advanced, optional)", "6.27"])]),
    7: dict(lookat=[["7.1", "7.2"], [], ["7.17", "7.19"]],
            inclass=[["7.4", "7.8"], ["7.9"], ["7.20", "7.33", "7.36"]],
            hw=[("HW08", ["7.3", "7.10", "7.14"]),
                ("HW09", ["7.27", "7.31", "7.34", "7.38 (optional challenge)",
                          "7.46 (Noether)"])]),
    8: dict(lookat=[["8.6", "8.8"], ["8.19", "8.20"], ["8.28", "8.32"]],
            inclass=[["8.3", "8.10"], ["8.12", "8.21"], ["8.29", "8.33", "8.34"]],
            hw=[("HW09", ["8.2"]),
                ("HW10", ["8.9", "8.13", "8.14 (optional challenge)", "8.15", "8.18", "8.23"]),
                ("HW11", ["8.27 (optional challenge)", "8.33"])]),
    9: dict(lookat=[["9.1"], ["9.7"], ["9.12", "9.13"]],
            inclass=[["9.2"], ["9.8"], ["9.14", "9.16", "9.19"]],
            hw=[("HW11", ["9.3", "9.9", "9.15", "9.17", "9.28", "9.29 (optional challenge)"])]),
    11: dict(lookat=[["11.1"], [], []],
             inclass=[["11.2", "11.3"], ["11.5"], []],
             hw=[("HW12", ["11.6", "11.9", "11.12", "11.14", "11.17 (optional)"])]),
    12: dict(lookat=[[], []],
             inclass=[["driven-pendulum notebook: period exploration, Poincare section"],
                      ["bifurcation diagram, Lyapunov exponent"]],
             hw=[("HW13", ["12.1", "12.6", "12.7", "12.8-12.10 (optional)",
                           "12.13 + parts (b), (c)"])]),
}


def write_chapter_problems(wb, md_path):
    """One sheet row per chapter-day, plus a markdown file for the slides."""
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    wrap = Alignment(wrap_text=True, vertical="top")
    ws = wb.create_sheet("Chapter Problems")
    ws.append(["Chapter", "Day", "Look at (PCCI source)", "In class", "Homework"])
    for cell in ws[1]:
        cell.font = header_font; cell.fill = header_fill
    lines = ["# Chapter problem lists (Taylor, Classical Mechanics 2005)", "",
             "Three streams per chapter, in Will Raven's F2025 format: **Look at** =",
             "study before class (solutions posted; our PCCIs are one of these per",
             "day), **In class** = worked together, **Homework** = the graded set.",
             "Generated by `make_fall2026_calendar.py` from `CHAPTER_PROBLEMS`; the",
             "chapter-opening slide is these three columns.", ""]
    for ch, d in CHAPTER_PROBLEMS.items():
        ndays = max(len(d["lookat"]), len(d["inclass"]))
        hw_lines = [f"{name}: " + ", ".join(items) for name, items in d["hw"]]
        for i in range(ndays):
            la = ", ".join(d["lookat"][i]) if i < len(d["lookat"]) else ""
            ic = ", ".join(d["inclass"][i]) if i < len(d["inclass"]) else ""
            ws.append([ch if i == 0 else None, f"Day {i + 1}", la, ic,
                       "\n".join(hw_lines) if i == 0 else None])
        lines += [f"## Chapter {ch}", "",
                  "| Day | Look at | In class |", "| --- | --- | --- |"]
        for i in range(ndays):
            la = ", ".join(d["lookat"][i]) if i < len(d["lookat"]) else "--"
            ic = ", ".join(d["inclass"][i]) if i < len(d["inclass"]) else "--"
            lines.append(f"| {i + 1} | {la or '--'} | {ic or '--'} |")
        lines += ["", "Homework: " + "; ".join(hw_lines), ""]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    for j, w in enumerate([9, 7, 30, 44, 60]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = w
    Path(md_path).write_text("\n".join(lines))

# ---------------------------------------------------------------- build
def build(outpath):
    days = list(class_days())
    n = len(days)
    assert len(CONTENT) + len(EXAMS) == n, (
        f"{len(CONTENT)} content + {len(EXAMS)} exams for {n} class meetings")
    assert all(days[i].weekday() in (0, 2, 4) for i in EXAMS), (
        "exam not on a class day")
    exam_days = {days[i]: EXAMS[i] for i in EXAMS}
    assert all(d in days for d in PCCI), "PCCI assigned to a non-class day"
    assert not any(d in exam_days for d in PCCI), "PCCI on an exam day"

    # ---- HW guards (the lens-grid rule: mechanical, hard-fail)
    hw_due = {}
    seen = set()
    for hw, due, through, chapters, covers, problems in HWS:
        assert hw not in seen, f"duplicate HW{hw:02d}"
        seen.add(hw)
        assert due > through, (
            f"HW{hw:02d} due {due} but draws on class {through}")
        assert through in days, f"HW{hw:02d} covers_through is not a class day"
        assert due not in NO_CLASS, f"HW{hw:02d} due on a no-class day"
        assert due not in exam_days, f"HW{hw:02d} due on an exam day"
        assert due.weekday() < 5, f"HW{hw:02d} due on a weekend"
        for ed, (label, ex_ch) in exam_days.items():
            if chapters & ex_ch:
                assert due < ed, (
                    f"HW{hw:02d} (Ch {sorted(chapters)}) due {due}, after "
                    f"{label} on {ed}")
        hw_due[hw] = due
    assert sorted(hw_due) == list(range(1, 14)), "expected HW01..HW13"
    assert all(hw_due[i] < hw_due[i + 1] for i in range(1, 13)), (
        "HW due dates out of order")

    # rows: one per class meeting; insert break markers and off-day HW rows
    rows = []      # (week, class_no, date, topic, reading, prereq, pcci, hw, exam)
    week_no = 0
    last_week = None
    class_no = 0
    content_i = 0
    breaks_seen = set()
    due_on = {}
    for hw, due in hw_due.items():
        due_on.setdefault(due, []).append(f"HW{hw:02d}")
    for d, label in EXTRA_DUE.items():
        due_on.setdefault(d, []).append(label)
    for slot_i, d in enumerate(days):
        iso_week = d.isocalendar()[1]
        if iso_week != last_week:
            week_no += 1
            last_week = iso_week
        for bd, why in NO_CLASS.items():
            if bd not in breaks_seen and bd < d:
                breaks_seen.add(bd)
                rows.append((None, None, bd, f"No class - {why}",
                             "", "", "", "", ""))
        if slot_i in EXAMS:
            label, _ = EXAMS[slot_i]
            topic, reading, prereq, exam = label, "", "", label
        else:
            (topic, reading, prereq), exam = CONTENT[content_i], ""
            content_i += 1
        hw = "; ".join(due_on.pop(d, []))
        class_no += 1
        rows.append((week_no, class_no, d, topic, reading, prereq,
                     PCCI.get(d, ""), hw, exam))
    for bd, why in NO_CLASS.items():
        if bd not in breaks_seen:
            rows.append((None, None, bd, f"No class - {why}",
                         "", "", "", "", ""))
    # HW due on a non-class day (HW13 in reading period) gets its own row.
    for d, labels in due_on.items():
        rows.append((None, None, d, "Reading period", "", "", "",
                     "; ".join(labels), ""))
    rows.sort(key=lambda r: r[2])
    rows.append((None, None, date(2026, 12, 19),
                 "Final exam period Dec 19-22 (self-scheduled)",
                 "", "", "", "",
                 "Final (Ch 12 required + optional grade-replacement)"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    break_fill = PatternFill("solid", fgColor="FCE4D6")
    exam_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["Week", "Class", "Date", "Topics", "Reading Due",
               "Prerequisites", "PCCI", "HW Due", "Exams"]
    ncols = len(headers)
    last_col = openpyxl.utils.get_column_letter(ncols)
    # Single canonical block, written as TWO CHUNKS (header row repeated),
    # split at fall break: the Moodle Page embeds chunk 1's range until
    # fall break, then chunk 2's (one URL-parameter edit; see
    # MoodleBuildSpec.md). The ranges are printed on every run.
    chunk2_start = date(2026, 10, 14)  # first class after autumn recess

    def web_header(r):
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=j, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = wrap

    web_header(1)
    r, split_row = 1, None
    for (wk, cn, d, topic, reading, prereq, pcci, hw, exam) in rows:
        if split_row is None and d >= chunk2_start:
            r += 1
            web_header(r)
            split_row = r
        r += 1
        vals = [wk, cn, d.strftime("%a %b %-d"), topic, reading, prereq,
                pcci, hw, exam]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = border
            cell.alignment = wrap
            if cn is None:
                cell.fill = break_fill
            elif exam:
                cell.fill = exam_fill
    for j, w in enumerate([7, 7, 10, 34, 12, 40, 18, 10, 16]):
        col = openpyxl.utils.get_column_letter(1 + j)
        ws.column_dimensions[col].width = w
    print(f"Schedule embed ranges: chunk 1 = A1:{last_col}{split_row - 1}, "
          f"chunk 2 = A{split_row}:{last_col}{r}")

    # ------------------------------------------------ HW problem lists
    hws = wb.create_sheet("HW Problem Lists")
    note = ("All problems from Taylor, Classical Mechanics (2005). Due "
            "Wednesdays at 10:00 PM on Moodle unless the Due column says "
            "otherwise (the exceptions sit right before an exam or a "
            "break). Custom problems are written out in full on the "
            "Moodle assignment page.")
    hws.append([note])
    hws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    hws.cell(row=1, column=1).alignment = wrap
    hws.row_dimensions[1].height = 45
    hws.append(["HW", "Due", "Covers", "Chapters", "Problems"])
    for cell in hws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for i, (hw, due, through, chapters, covers, problems) in enumerate(
            HWS, start=3):
        vals = [f"HW{hw:02d}", due.strftime("%a %b %-d"), covers,
                ", ".join(str(c) for c in sorted(chapters)), problems]
        for j, v in enumerate(vals, start=1):
            cell = hws.cell(row=i, column=j, value=v)
            cell.border = border
            cell.alignment = wrap
    for j, w in enumerate([8, 11, 40, 10, 60]):
        hws.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = w

    # -------------------------------------------------- grade categories
    gc = wb.create_sheet("Grade Categories")
    gc.append(["Category", "Number", "Drop", "Points Each", "Total Points"])
    for cell in gc[1]:
        cell.font = header_font
        cell.fill = header_fill
    # Revised 2026-09-04 (Michael): each exam is three problems, one per
    # chapter, and the final's required problem is one more of the same
    # kind -- so every exam-type problem is worth exactly 60 points
    # (3 x 60 = 180 per exam), and the remainder is split as participation
    # 112 + homework 288. Course total must be exactly 1000.
    cats = [
        ("Attendance/participation (PCCIs)", 39, 4, 3.2),
        ("Weekly Homework", 13, 1, 24),
        ("Exams", 3, 0, 180),
        ("Final exam (Ch 12 problem)", 1, 0, 60),
    ]
    total = sum((num - drop) * pts for _, num, drop, pts in cats)
    assert total == 1000, f"grade categories sum to {total}, not 1000"
    for i, (name, num, drop, pts) in enumerate(cats, start=2):
        gc.append([name, num, drop, pts, f"=(B{i}-C{i})*D{i}"])
    gc.append(["Total", None, None, None, f"=SUM(E2:E{1 + len(cats)})"])
    for j, w in enumerate([34, 9, 7, 12, 13]):
        gc.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = w

    write_chapter_problems(wb, Path(outpath).with_name("ChapterProblemLists.md"))

    wb.save(outpath)
    print(f"wrote {outpath}: {len(rows)} schedule rows, {n} class meetings, "
          f"{len(HWS)} HWs, grade total {total:g}")

if __name__ == "__main__":
    build(sys.argv[1])
